from pathlib import Path

from openpyxl import Workbook, load_workbook

from pipeline.entities import Candidate
from utils.logger import logger


class ExcelExporter:

    OUTPUT_DIR = Path("output")
    FILE_NAME = "resumes.xlsx"

    FILE_PATH = OUTPUT_DIR / FILE_NAME

    HEADERS = [
        "Name",
        "Email",
        "Phone",
        "LinkedIn",
        "GitHub",
        "Portfolio",
        "Skills",
        "Education",
        "Experience",
        "Projects",
        "Certifications",
    ]

    @classmethod
    def export(cls, candidate: Candidate) -> None:

        logger.info("Exporting candidate to Excel.")

        workbook, sheet = cls._create_workbook()

        row = cls._candidate_to_row(candidate)

        sheet.append(row)

        workbook.save(cls.FILE_PATH)

        logger.info("Candidate exported successfully.")

    @classmethod
    def _create_workbook(cls):

        cls.OUTPUT_DIR.mkdir(exist_ok=True)

        if cls.FILE_PATH.exists():

            workbook = load_workbook(cls.FILE_PATH)

            sheet = workbook.active

            return workbook, sheet

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Candidates"

        sheet.append(cls.HEADERS)

        return workbook, sheet
    
    @classmethod
    def _candidate_to_row(cls, candidate: Candidate):

        info = candidate.personal_info
        social = candidate.social_links

        return [

            info.name,

            info.email,

            info.phone,

            social.linkedin,

            social.github,

            social.portfolio,

            ", ".join(candidate.skills),

            cls._education(candidate),

            cls._experience(candidate),

            cls._projects(candidate),

            ", ".join(candidate.certifications),
        ]
    
    @staticmethod
    def _education(candidate):

        records = []

        for education in candidate.education:

            records.append(
                f"{education.degree} | "
                f"{education.institute} | "
                f"{education.duration} | "
                f"CGPA: {education.cgpa}"
            )

        return "\n".join(records)
    
    @staticmethod
    def _experience(candidate):

        records = []

        for experience in candidate.experience:

            records.append(
                f"{experience.designation} @ "
                f"{experience.company} "
                f"({experience.duration})"
            )

        return "\n".join(records)
    
    @staticmethod
    def _projects(candidate):

        records = []

        for project in candidate.projects:

            records.append(project.title)

        return "\n".join(records)